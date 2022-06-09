from pathlib import Path
import openpyxl as xl


def grant(args):
    # open excel file
    path = Path(__file__).parents[1]
    excel_file = 'rules.xlsx'
    wb = xl.load_workbook(excel_file)
    rules = wb.active

    # args check
    output = ''
    if len(args) < 3:
        output = 'Команда Grant должна содержать не менее 3 аргументов!'
    else:
        users = [u.value for u in rules['A'][1:]]
        symbols = [s.value for s in rules['1'][1:]]
        if args[0] in users and args[1] in users:
            for arg in args[2:]:
                if arg not in symbols:
                    output = 'Команда содержит несуществующий символ!'
                    break
        else:
            output = 'Команда содержит несуществующего пользователя!'

        if not output:
            subj1 = None
            subj2 = None
            for u in rules['A'][1:]:
                if u.value == args[0]:
                    subj1 = u
                if u.value == args[1]:
                    subj2 = u

            for arg in args[2:]:
                for s in rules['1'][1:]:
                    if s.value == arg:
                        if rules[subj1.row][s.column-1].value == 1:
                            rules[subj2.row][s.column-1].value = 1
                        else:
                            output = 'Пользователь не может передать права на один из объектов!'
                            return output
            output = 'Пользователь %s передал права на объект(ы) пользователю %s' % (subj1.value, subj2.value)
    wb.save(excel_file)
    return output


def remove(args):
    # open excel file
    path = Path(__file__).parents[1]
    excel_file = 'rules.xlsx'
    wb = xl.load_workbook(excel_file)
    rules = wb.active

    # args check
    output = ''
    users = [u.value for u in rules['A'][1:]]
    symbols = [s.value for s in rules['1'][1:]]
    if args[0] in users:
        for arg in args[1:]:
            if arg not in symbols:
                output = 'Команда содержит несуществующий символ!'
                break
    else:
        output = 'Команда содержит несуществующего пользователя!'

    if not output:
        for u in rules['A'][1:]:
            if u.value == args[0]:
                for arg in args[1:]:
                    for s in rules['1'][1:]:
                        if s.value == arg:
                            rules[u.row][s.column-1].value = 0
                            output = 'Изъяты права пользователя %s на объект(ы)' % u.value
    wb.save(excel_file)
    return output


def create(args):
    # open excel file
    path = Path(__file__).parents[1]
    excel_file = 'rules.xlsx'
    wb = xl.load_workbook(excel_file)
    rules = wb.active
    output = ''

    # args check
    users = [u.value for u in rules['A'][1:]]
    symbols = [s.value for s in rules['1'][1:]]
    if args[0] in users:
        for arg in args[1:]:
            if len(arg) == 1:
                if arg in symbols:
                    output = 'Команда Сreate не может давать права на существующий символ!'
                    break
            else:
                output = 'Команда содержит непраильный символ!'
                break
    else:
        row = rules['A'][-1].row + 1
        last = rules.cell(row=row, column=1, value=args[0])
        for i in range(len(symbols)):
            rules.cell(row = row, column = i + 2, value = 0)





    if not output:
        for u in rules['A'][1:]:
            if u.value == args[0]:
                for arg in args[1:]:
                    col = rules['1'][-1].column + 1
                    last = rules.cell(row=1, column=col, value=arg)
                    for cell in rules[last.column_letter][1:]:
                        if cell.row == u.row:
                            cell.value = 1
                        else:
                            cell.value = 0
        output = 'Новые(ый) объект(ы) были успешно созданы'
    wb.save(excel_file)
    return output
